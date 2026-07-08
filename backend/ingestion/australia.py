"""
Australia Consolidated Sanctions List ingester.

Source:
  https://www.dfat.gov.au/sites/default/files/Australian_Sanctions_Consolidated_List.xlsx

Notes:
- DFAT publishes this as an XLSX dataset (not a JSON API).
- We parse the sheet using header-based mapping so minor column order changes
  do not break ingestion.
"""

from __future__ import annotations

import hashlib
import io
import logging
import os
import re
from typing import Any

import httpx
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database.repository import SanctionRepository, SyncStateRepository
from ingestion.base import BaseIngester

log = logging.getLogger(__name__)

_XLSX_URL = "https://www.dfat.gov.au/sites/default/files/Australian_Sanctions_Consolidated_List.xlsx"
_TIMEOUT_SECONDS = 180.0
_MAX_RETRIES = 2
_BACKOFF_SECONDS = 1.0
_LOCAL_XLSX_ENV = "AUSTRALIA_XLSX_PATH"


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _split_multi(value: str | None) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[;\n]+", value)
    out: list[str] = []
    for part in parts:
        token = part.strip()
        if token and token not in out:
            out.append(token)
    return out


_HEADER_CANDIDATES: dict[str, set[str]] = {
    "list_id": {
        "reference",
        "ref",
        "referencenumber",
        "listingreference",
        "designationreference",
        "id",
    },
    "name": {
        "name",
        "nameofindividualorentity",
        "individualorentity",
        "designatedpersonorentity",
        "personorentityname",
    },
    "entity_type": {
        "type",
        "listingtype",
        "entitytype",
        "personorentitytype",
    },
    "aliases": {
        "alias",
        "aliases",
        "aka",
        "alsoKnownAs",
        "othernames",
        "namealias",
    },
    "country": {
        "country",
        "citizenship",
        "nationality",
        "countryofcitizenship",
        "countries",
    },
    "address": {
        "address",
        "addresses",
        "residentialaddress",
        "businessaddress",
    },
    "programs": {
        "sanctionsregime",
        "regime",
        "listinginformation",
        "listingcriteria",
        "framework",
        "committee",
        "committeename",
    },
    "remarks": {
        "additionalinformation",
        "otherinformation",
        "comments",
        "remarks",
        "details",
    },
    "dob": {
        "dateofbirth",
        "dob",
    },
    "pob": {
        "placeofbirth",
        "pob",
    },
}


def _match_field(header: str) -> str | None:
    norm = _norm(header)
    if not norm:
        return None

    for field, candidates in _HEADER_CANDIDATES.items():
        if norm in {_norm(c) for c in candidates}:
            return field

    # Fuzzy contains checks for robustness.
    if "name" in norm and ("entity" in norm or "individual" in norm):
        return "name"
    if "alias" in norm:
        return "aliases"
    if "citizenship" in norm or "nationality" in norm:
        return "country"
    if "address" in norm:
        return "address"
    if "regime" in norm or "committee" in norm:
        return "programs"
    if "additional" in norm or "remark" in norm or "otherinformation" in norm:
        return "remarks"
    if "dateofbirth" in norm:
        return "dob"
    if "placeofbirth" in norm:
        return "pob"
    if "reference" in norm:
        return "list_id"
    if norm == "type":
        return "entity_type"
    return None


def _parse_entity_type(raw: str | None) -> str:
    value = (raw or "").strip().lower()
    if not value:
        return "company"
    if "individual" in value or "person" in value:
        return "individual"
    if "vessel" in value or "ship" in value:
        return "vessel"
    if "aircraft" in value or "plane" in value:
        return "aircraft"
    return "company"


def _to_country_code(raw: str | None) -> str | None:
    if not raw:
        return None
    value = raw.strip().upper()
    if len(value) == 2 or len(value) == 3:
        return value
    return None


class AustraliaIngester(BaseIngester):
    list_source = "AUSTRALIA"

    def __init__(self):
        self._cached_bytes: bytes | None = None
        self._cached_publication_id: int | None = None

    def _download_bytes(self) -> bytes:
        local_path = os.getenv(_LOCAL_XLSX_ENV, "").strip()
        if local_path:
            with open(local_path, "rb") as f:
                return f.read()

        timeout = httpx.Timeout(_TIMEOUT_SECONDS, connect=20.0)
        last_error: Exception | None = None

        for attempt in range(_MAX_RETRIES + 1):
            try:
                response = httpx.get(_XLSX_URL, timeout=timeout, follow_redirects=True)
                response.raise_for_status()
                return response.content
            except Exception as exc:
                last_error = exc
                if attempt >= _MAX_RETRIES:
                    break

                delay_seconds = _BACKOFF_SECONDS * (attempt + 1)
                import time

                time.sleep(delay_seconds)

        raise RuntimeError(f"Failed to download Australia sanctions XLSX: {last_error}")

    def _compute_publication_id(self, payload: bytes) -> int:
        digest = hashlib.sha256(payload).hexdigest()[:15]
        return int(digest, 16)

    def needs_update(self, db: Session) -> bool:
        payload = self._download_bytes()
        publication_id = self._compute_publication_id(payload)

        self._cached_bytes = payload
        self._cached_publication_id = publication_id

        state = SyncStateRepository(db).get(self.list_source)
        if not state or not state.last_publication_id:
            return True

        return publication_id != state.last_publication_id

    def _find_header_map(self, rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
        best_idx = -1
        best_map: dict[str, int] = {}

        for idx, row in enumerate(rows):
            candidate_map: dict[str, int] = {}
            for col_idx, cell in enumerate(row):
                field = _match_field(str(cell) if cell is not None else "")
                if field and field not in candidate_map:
                    candidate_map[field] = col_idx

            score = len(candidate_map)
            has_name = "name" in candidate_map
            if has_name and score >= 2 and score > len(best_map):
                best_idx = idx
                best_map = candidate_map

        if best_idx < 0:
            raise RuntimeError("Could not detect header row in Australia sanctions XLSX.")

        return best_idx, best_map

    def _parse_records(self, payload: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        sheet = workbook.active

        sample_rows = list(sheet.iter_rows(min_row=1, max_row=30, values_only=True))
        header_idx, header_map = self._find_header_map(sample_rows)

        records: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()

        for absolute_idx, row in enumerate(
            sheet.iter_rows(min_row=header_idx + 2, values_only=True),
            start=header_idx + 2,
        ):
            if not row:
                continue

            def pick(field: str) -> str | None:
                col = header_map.get(field)
                if col is None or col >= len(row):
                    return None
                return _text(row[col])

            name = pick("name")
            if not name:
                continue

            list_id = pick("list_id") or f"AU-{absolute_idx}"
            aliases = _split_multi(pick("aliases"))
            country_raw = pick("country")
            country_code = _to_country_code(country_raw)
            address = pick("address")
            programs = _split_multi(pick("programs"))

            remarks_parts = [
                pick("remarks"),
                f"DOB: {pick('dob')}" if pick("dob") else None,
                f"POB: {pick('pob')}" if pick("pob") else None,
                f"Country/Nationality: {country_raw}" if country_raw else None,
            ]
            remarks = "; ".join(part for part in remarks_parts if part)
            remarks = remarks if remarks else None

            entity_type = _parse_entity_type(pick("entity_type"))

            dedupe_key = (name.casefold(), str(list_id).casefold())
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)

            records.append(
                {
                    "name": name,
                    "aliases": aliases,
                    "country": country_code,
                    "list_source": self.list_source,
                    "list_id": str(list_id),
                    "entity_type": entity_type,
                    "address": address,
                    "programs": programs,
                    "remarks": remarks,
                }
            )

        return records

    def ingest(self, db: Session) -> int:
        sync_repo = SyncStateRepository(db)
        sanction_repo = SanctionRepository(db)

        payload = self._cached_bytes or self._download_bytes()
        publication_id = self._cached_publication_id or self._compute_publication_id(payload)

        try:
            records = self._parse_records(payload)
        except Exception as exc:
            sync_repo.upsert(self.list_source, publication_id=publication_id, status="failed")
            raise RuntimeError(f"Failed to parse Australia sanctions XLSX: {exc}") from exc

        if not records:
            sync_repo.upsert(self.list_source, publication_id=publication_id, status="failed")
            raise RuntimeError("Australia sanctions XLSX parsed but produced zero records.")

        sanction_repo.clear_list(self.list_source)
        sanction_repo.bulk_add(records)

        sync_repo.upsert(
            self.list_source,
            publication_id=publication_id,
            entity_count=len(records),
            status="ok",
        )

        log.info("Australia ingest complete: %s records loaded.", len(records))
        return len(records)





