"""
Generates a colour-coded Excel workbook for a screening run.

Sheets:
  1. Summary   — run metadata and counts
  2. Results   — all vendors with status colour coding
  3. Flagged   — flagged and review-needed vendors only
"""
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from database.models import ScreeningRun, MatchStatus

# ── Colour palette ────────────────────────────────────────────────────────────
_RED    = PatternFill("solid", fgColor="FFCCCC")
_YELLOW = PatternFill("solid", fgColor="FFF2CC")
_GREEN  = PatternFill("solid", fgColor="CCFFCC")
_GREY   = PatternFill("solid", fgColor="D9D9D9")
_DARK   = PatternFill("solid", fgColor="1F4E79")

_WHITE_BOLD = Font(bold=True, color="FFFFFF")
_BOLD       = Font(bold=True)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

_STATUS_FILL = {
    MatchStatus.FLAGGED: _RED,
    MatchStatus.REVIEW:  _YELLOW,
    MatchStatus.CLEAR:   _GREEN,
}
_STATUS_LABEL = {
    MatchStatus.FLAGGED: "FLAGGED",
    MatchStatus.REVIEW:  "REVIEW NEEDED",
    MatchStatus.CLEAR:   "CLEAR",
}


def _header_row(ws, columns: list[str], row: int = 1):
    for col_idx, title in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.fill   = _DARK
        cell.font   = _WHITE_BOLD
        cell.alignment = _CENTER
        cell.border = _THIN


def _autofit(ws, min_width=12, max_width=50):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(length + 2, min_width), max_width
        )


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def _build_summary_sheet(ws, run: ScreeningRun):
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    def row(label, value, bold=False):
        r = ws.max_row + 1
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=value)
        a.font = _BOLD if bold else Font()
        a.fill = _GREY
        a.alignment = _LEFT
        b.alignment = _LEFT
        a.border = _THIN
        b.border = _THIN

    ws.append([])
    ws.cell(row=1, column=1, value="TRADE SANCTIONS SCREENING REPORT").font = Font(bold=True, size=14)

    ws.append([])
    row("Report ID",          f"SCR-{run.id:06d}", bold=True)
    row("Customer",           run.customer_name)
    row("Generated At",       datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
    row("Screening Duration", f"{run.elapsed_seconds:.2f} seconds" if run.elapsed_seconds else "—")
    row("Run Status",         run.status.value.upper())

    ws.append([])
    flagged = sum(1 for r in run.results if r.status == MatchStatus.FLAGGED)
    review  = sum(1 for r in run.results if r.status == MatchStatus.REVIEW)
    clear   = sum(1 for r in run.results if r.status == MatchStatus.CLEAR)

    row("Vendors Screened",   len(run.results), bold=True)
    row("Flagged",            flagged)
    row("Review Needed",      review)
    row("Clear",              clear)

    if run.ai_summary:
        ws.append([])
        ws.cell(row=ws.max_row + 1, column=1, value="AI Narrative").font = _BOLD
        narrative_cell = ws.cell(row=ws.max_row + 1, column=1, value=run.ai_summary)
        narrative_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(
            start_row=narrative_cell.row, start_column=1,
            end_row=narrative_cell.row,   end_column=2
        )
        ws.row_dimensions[narrative_cell.row].height = 200


# ── Sheet 2: All Results ──────────────────────────────────────────────────────

_RESULT_COLS = [
    "Vendor Name", "Status", "Match Score (%)", "Matched Entity",
    "Sanctions List", "Match Type", "Supplier Tier", "AI Reasoning",
]


def _build_results_sheet(ws, run: ScreeningRun, results=None):
    ws.title = "All Results"
    _header_row(ws, _RESULT_COLS)

    for r in (results or run.results):
        row_data = [
            r.vendor_name,
            _STATUS_LABEL.get(r.status, r.status.value),
            round(r.match_score, 1) if r.match_score else None,
            r.matched_name,
            r.list_source,
            r.match_type,
            r.tier,
            r.ai_reasoning,
        ]
        ws.append(row_data)
        fill = _STATUS_FILL.get(r.status, _GREEN)
        for col_idx in range(1, len(_RESULT_COLS) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill      = fill
            cell.alignment = _LEFT
            cell.border    = _THIN

    _autofit(ws)
    ws.freeze_panes = "A2"


# ── Sheet 3: Flagged Only ─────────────────────────────────────────────────────

def _build_flagged_sheet(ws, run: ScreeningRun):
    ws.title = "Flagged & Review"
    at_risk = [
        r for r in run.results
        if r.status in (MatchStatus.FLAGGED, MatchStatus.REVIEW)
    ]
    if not at_risk:
        ws.append(["No flagged or review-needed vendors in this run."])
        return
    _build_results_sheet(ws, run, results=at_risk)
    ws.title = "Flagged & Review"   # reset after _build_results_sheet renames it


# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel(run: ScreeningRun) -> bytes:
    """Return Excel workbook as bytes (ready to stream via HTTP response)."""
    wb = Workbook()

    ws_summary = wb.active
    _build_summary_sheet(ws_summary, run)

    ws_results = wb.create_sheet()
    _build_results_sheet(ws_results, run)

    ws_flagged = wb.create_sheet()
    _build_flagged_sheet(ws_flagged, run)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
